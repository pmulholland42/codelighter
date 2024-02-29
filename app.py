import tkinter as tk
from tkinter import filedialog, Listbox, Spinbox, IntVar, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import sys

def clear_widgets():
  # Clear the window and display the list of sheet names
  for widget in root.winfo_children():
      widget.destroy()

def append_highlighted_to_filename(filename):
    # Split the filename on the last dot to get the base name and extension
    parts = filename.rsplit('.', 1)
    
    # If the filename contains an extension, insert '-highlighted' before the extension
    if len(parts) > 1:
        new_filename = f"{parts[0]}-highlighted.{parts[1]}"
    else:
        # If there's no extension, just append '-highlighted'
        new_filename = f"{filename}-highlighted"
    
    return new_filename

class AccordionItem(tk.Frame):
    def __init__(self, master, title="", content=[]):
        super().__init__(master)
        self.canvas = None
        self.scrollable_frame = None
        self.is_expanded = False
        self.scrollbar = None

        self.title_button = tk.Button(self, text=title, command=self.toggle)
        self.title_button.pack(fill=tk.X)

        self.set_content(content)

    def set_content(self, content):
        if self.canvas is not None:
            self.canvas.destroy()
        if self.scrollbar is not None:
            self.scrollbar.destroy()

        self.canvas = tk.Canvas(self)
        self.scrollbar = tk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(
                scrollregion=self.canvas.bbox("all")
            )
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        for content_item in content:
            content_label = tk.Label(self.scrollable_frame, text=content_item)
            content_label.pack()

    def toggle(self):
        if self.is_expanded:
            self.canvas.pack_forget()
            self.scrollbar.pack_forget()
        else:
            self.canvas.pack(side="left", fill="both", expand=True)
            self.scrollbar.pack(side="right", fill="y")
        self.is_expanded = not self.is_expanded

def create_accordion_item(master, title, content):
    item = AccordionItem(master, title=title, content=content)
    item.pack(fill=tk.X, padx=5, pady=2)


def select_xlsx_file():
    # Define the file types that should be allowed, in this case, .xlsx files
    filetypes = [('Excel files', '*.xlsx')]
    documents_path = os.path.expanduser('~/Documents')
    
    # Open the file dialog and store the selected file path
    filepath = filedialog.askopenfilename(title='Open a file', initialdir=documents_path, filetypes=filetypes)

    if filepath:
        print(f"Selected file: {filepath}")
        # Try to open the workbook and list sheet names
        try:
            filename = os.path.basename(filepath)
            wb = load_workbook(filename=filepath)
            sheet_names = wb.sheetnames
            if (len(sheet_names) > 1):
                display_sheet_selection(wb, filename)
            else:
                display_max_row_selection(wb, sheet_names[0], filename, False)
        except Exception as e:
            print("Error reading file:", e)
            display_file_selection("Some error")
    else:
        print("No file selected.")

def display_sheet_selection(wb, filename):
    clear_widgets()
    sheet_names = wb.sheetnames
    # Function to handle sheet selection and move to the next action
    def on_continue():
        selected_sheet = sheet_names[listbox.curselection()[0]]
        display_max_row_selection(wb, selected_sheet, filename, True)
        print(f"Selected sheet: {selected_sheet}")
        # Proceed with your next action here, for example, opening the selected sheet
        # This is where you would add the logic to manipulate or display the selected sheet
    
    label = tk.Label(root, text="Select a sheet:")
    label.pack()
    
    listbox = Listbox(root)
    listbox.pack(fill=tk.BOTH, expand=True)
    
    for name in sheet_names:
        listbox.insert(tk.END, name)
        root.update()

    continue_button = tk.Button(root, text="Continue", command=on_continue)
    continue_button.pack(side=tk.RIGHT)

    back_button = tk.Button(root, text="Back", command=display_file_selection)
    back_button.pack(side=tk.RIGHT)  
    root.update()
  

def display_file_selection(error_text = ""):
  clear_widgets()
  label = tk.Label(root, text="To begin, select an Excel spreadsheet with your coding data.")
  label.pack()
    # Create a button that will open the file dialog when clicked
  select_file_btn = tk.Button(root, text='Select File', command=select_xlsx_file)
  select_file_btn.pack(expand=True)
  if (len(error_text) > 0):
      error_label = tk.Label(root, text=error_text, fg="red")
      error_label.pack()

def display_max_row_selection(wb, sheet_name, filename, from_sheet_select):
    clear_widgets()
    sheet = wb[sheet_name]
    max_row = sheet.max_row

    label = tk.Label(root, text="Enter the maximum row number that you want to be highlighted:")
    label.pack()
    max_row_var = IntVar(root)
    max_row_var.set(1)  # Default value
    spinbox = Spinbox(root, from_=1, to=max_row, textvariable=max_row_var, width=10)
    spinbox.pack()
    root.update()

    back_command = lambda: print()
    if from_sheet_select:
        back_command = lambda: display_sheet_selection(wb, filename)
    else:
        back_command = display_file_selection

    confirm_button = tk.Button(root, text="Continue", command=lambda: highlight_rows(wb, sheet_name, max_row_var.get(), filename))
    confirm_button.pack(side=tk.RIGHT)  

    back_button = tk.Button(root, text="Back", command=back_command)
    back_button.pack(side=tk.RIGHT)   
    
    root.update()

def highlight_rows(wb, sheet_name, max_row, filename):
  fill =  PatternFill("solid", fgColor="5B9BD5")
  ws = wb[sheet_name]
  max_col = 200

  column_errors = set()
  success_cols = set()
  highlighted_count = 0
  total_count = 0

  for col in ws.iter_cols(min_row=1, max_row=max_row, min_col=0, max_col=max_col):
      column_name = col[0].value
      if isinstance(column_name, str) and column_name.endswith('_1'):
          column_base_name = '_'.join(column_name.split('_')[:-1])
          second_column = {}
          second_column_name = ""

          for second_col in ws.iter_cols(min_row=1, max_row=max_row, min_col=0, max_col=max_col):
              second_column_name = second_col[0].value
              if isinstance(second_column_name, str) and second_column_name.endswith('_2'):
                  second_column_base_name = '_'.join(second_column_name.split('_')[:-1])
                  if second_column_base_name == column_base_name:
                      second_column = second_col
                      break
          
          for index, cell in enumerate(col):
              if cell.value == 0 or cell.value == 1:
                  try:
                      if (cell.value != second_column[index].value):
                          cell.fill = fill
                          highlighted_count = highlighted_count + 1
                      total_count = total_count + 1
                      success_cols.add('Compared ' + column_name + ' to ' + second_column_name)
                  except:
                      column_errors.add("Could not read value in " + column_name)
                      print(f'Exception getting second column value for {column_name} {index}')
  
  display_confirmation_window(wb, success_cols, column_errors, highlighted_count, total_count, filename)

def display_confirmation_window(wb, success_cols, errors, highlighted_count, total_count, filename):
    clear_widgets()        

    count_label = tk.Label(root, text="Done highlighting - here are the results:")
    count_label.pack()

    create_accordion_item(root, f"Compared {len(success_cols)} columns (click to toggle details)", success_cols)

    count_label = tk.Label(root, text=f"Checked {total_count} cells")
    count_label.pack()

    highlighted_count_label = tk.Label(root, text=f"Highlighted {highlighted_count} cells with differences")
    highlighted_count_label.pack()

    error_color = "red"
    if (len(errors) == 0):
      error_color = "black"
      error_label = tk.Label(root, text=f"{len(errors)} errors", fg=error_color)
      error_label.pack()
    else:
      create_accordion_item(root, f"{len(errors)} errors (click to toggle details)", errors)


    save_btn = tk.Button(root, text="Save Output File", command=lambda: save_file(wb, filename))
    save_btn.pack(side=tk.RIGHT)
        

def save_file(wb, filename):
    # Specify the default extension and file types
    filetypes = [('Excel files', '*.xlsx'), ('All files', '*.*')]
    initial_filename = append_highlighted_to_filename(filename)
    # Open the save file dialog
    output_file_name = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=filetypes, initialfile=initial_filename)
    if output_file_name:
        try:
            # Save the workbook to the chosen path
            wb.save(output_file_name)
            messagebox.showinfo("Success", "File saved successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save file: {e}")

    start_over_btn = tk.Button(root, text="Highlight another file", command=display_file_selection)
    start_over_btn.pack(side=tk.RIGHT)  

    exit_btn = tk.Button(root, text="Exit", command=sys.exit)
    exit_btn.pack(side=tk.RIGHT)  
  



# Create the main window
root = tk.Tk()
root.title('Select an .xlsx File')
root.geometry('400x400')  # Adjust the size as needed

display_file_selection()

# Start the Tkinter event loop
root.mainloop()
