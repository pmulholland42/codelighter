from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import random

input_file_name = input("Enter the file name without the extension: ")
sheet_name = str(input("Enter the name of the sheet: "))
max_row = input("Enter the max row: ")
max_col=200

print("Opening the file...")

try:
    wb = load_workbook(filename = input_file_name + '.xlsx')
except:
    print("Couldn't find '" + input_file_name + ".xlsx'. Make sure you run this program in the same folder as the excel file.")
    exit()

try:
    ws = wb[sheet_name]
except:
    print("Couldn't find a sheet named '" + sheet_name + "' in '" + input_file_name + ".xlsx'. Make sure the name is correct.")
    exit()

try:
    max_row = int(max_row)
except:
    print("Please enter a valid number of rows")
    exit()

print("Working...")


fill =  PatternFill("solid", fgColor="5B9BD5")

for col in ws.iter_cols(min_row=1, max_row=max_row, min_col=0, max_col=max_col):
    column_name = col[0].value
    if isinstance(column_name, str) and column_name.endswith('_1'):
        column_base_name = '_'.join(column_name.split('_')[:-1])
        second_column = {}

        for second_col in ws.iter_cols(min_row=1, max_row=max_row, min_col=0, max_col=max_col):
            second_column_name = second_col[0].value
            if isinstance(second_column_name, str) and second_column_name.endswith('_2'):
                second_column_base_name = '_'.join(second_column_name.split('_')[:-1])
                if second_column_base_name == column_base_name:
                    second_column = second_col
                    print('Comparing ' + column_name + ' to ' + second_column_name)
                    break

        
        
        for index, cell in enumerate(col):
            if cell.value == 0 or cell.value == 1:
                try:
                    if (cell.value != second_column[index].value):
                        cell.fill = fill
                except:
                    print(f'Exception getting second column value for {column_name} {index}')

print("Done! Saving...")
output_file_name = input_file_name + '-output.xlsx'

wb.save(output_file_name)
print('Saved output to ' + output_file_name)

print()

final_messages = ['I love you \u2764\uFE0F', 'Proud of you!', "Remember: you're beautiful", 'xoxoxoxo', 'I \u2764\uFE0F you', 'I appreciate you so much', 'u r cute', 'You are so sexy babe', 'You mean the world to me \u2764\uFE0F', 'Love you!', 'Love you xoxoxo', "Hope you're having a wonderful day :)"]
print(random.choice(final_messages))




